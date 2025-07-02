from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sentence_transformers import SentenceTransformer
from qdrant_client import QdrantClient
from qdrant_client.models import VectorParams, Distance, PointStruct
from docx import Document
import fitz  # PyMuPDF
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openai
from datetime import datetime
import pytz
import os

# === Cargar configuración de entorno y validar ===
def cargar_config():
    load_dotenv()
    config = {
        "OPENAI_API_KEY": os.getenv("OPENAI_API_KEY"),
        "QDRANT_URL": os.getenv("QDRANT_URL"),
        "QDRANT_API_KEY": os.getenv("QDRANT_API_KEY")
    }

    if not all(config.values()):
        raise EnvironmentError("Faltan variables de entorno requeridas.")
    
    return config

config = cargar_config()

# === Inicializar servicios ===
openai.api_key = config["OPENAI_API_KEY"]
model_embeddings = SentenceTransformer("sentence-transformers/all-MiniLM-L6-v2")

qdrant_client = QdrantClient(
    url=config["QDRANT_URL"],
    api_key=config["QDRANT_API_KEY"],
    timeout=120  # Aumentar timeout a 120 segundos
)

# === Constantes de la app ===
UPLOAD_FOLDER = "docs_upload"
COLLECTION_NAME = "documentos_legales_qdrant_jess"  # Nombre más genérico
CHUNK_SIZE = 800  # Aumentado para mejor contexto
OVERLAP_SIZE = 200  # Superposición entre chunks para mejor coherencia
MODEL_DIM = 384
EXCEL_PATH = "registro_chat.xlsx"
MIN_SIMILARITY_THRESHOLD = 0.3  # Umbral mínimo de similitud
BATCH_SIZE = 50  # Tamaño de lote para inserción en Qdrant
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-3.5-turbo")  # Modelo configurable

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Variable global para almacenar información del documento actual
documento_actual = {
    "tipo": None,
    "especialidad": None,
    "descripcion": None,
    "filename": None,
    "fecha_carga": None
}

# === FastAPI App ===
app = FastAPI(title="Chatbot Leyes")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# === Utilidades ===
def inicializar_qdrant():
    qdrant_client.recreate_collection(
        collection_name=COLLECTION_NAME,
        vectors_config=VectorParams(size=MODEL_DIM, distance=Distance.COSINE)
    )

# extrae texto de pdf y lo divide en fragmentos con superposición
def pdf_a_chunks(file_path: str, chunk_size: int = CHUNK_SIZE, overlap_size: int = OVERLAP_SIZE):
    texto = ""
    with fitz.open(file_path) as doc:
        for page in doc:
            page_text = page.get_text()
            if page_text.strip():  # Solo agregar páginas con contenido
                texto += page_text.strip() + "\n\n"

    texto = texto.strip()
    if not texto:
        raise ValueError("No se pudo extraer texto del PDF")
    
    # Detectar tipo de documento
    filename = os.path.basename(file_path)
    tipo_doc = detectar_tipo_documento(texto, filename)
    
    # Crear chunks con superposición para mejor coherencia contextual
    chunks = []
    start = 0
    
    while start < len(texto):
        end = start + chunk_size
        chunk = texto[start:end]
        
        # Si no es el último chunk, intentar cortar en un punto natural (punto, salto de línea)
        if end < len(texto):
            # Buscar el último punto o salto de línea en los últimos 100 caracteres
            natural_break = max(
                chunk.rfind('.', max(0, len(chunk) - 100)),
                chunk.rfind('\n', max(0, len(chunk) - 100))
            )
            if natural_break > 0:
                chunk = chunk[:natural_break + 1]
                end = start + len(chunk)
        
        chunks.append(chunk.strip())
        
        # Mover el inicio considerando la superposición
        if end >= len(texto):
            break
        start = end - overlap_size
    
    # Filtrar chunks vacíos o muy pequeños
    chunks = [chunk for chunk in chunks if len(chunk.strip()) > 50]
    
    vectores = model_embeddings.encode(chunks)
    return chunks, vectores, tipo_doc

def construir_prompt(contexto: str, pregunta: str, tipo_documento: dict, tiene_contexto_relevante: bool = True) -> str:
    if tiene_contexto_relevante:
        especialidad = tipo_documento.get('especialidad', 'Derecho General')
        descripcion = tipo_documento.get('descripcion', 'documento legal')
        tipo = tipo_documento.get('tipo', 'Documento Legal')
        
        return f"""Actúas como un JUEZ especializado en {especialidad} ecuatoriano.

INSTRUCCIONES OBLIGATORIAS:
Debes estructurar tu respuesta como una SENTENCIA JUDICIAL con el siguiente formato:

📅 **FECHA Y HORA:**
[Fecha y hora actual de la sentencia]

⚖️ **RAZÓN DE LA SENTENCIA:**
[Análisis jurídico detallado basado en el {tipo} proporcionado. Cita artículos específicos y fundamentos legales aplicables al caso presentado]

🏛️ **VEREDICTO:**
[CULPABLE/INOCENTE - con justificación legal específica basada en los artículos del {tipo}]

🏢 **LUGAR DE RECLUSIÓN:**
[Si es culpable: especificar centro penitenciario según gravedad del delito]
[Si es inocente: "Libertad inmediata del procesado"]

📋 **CONCLUSIÓN:**
[Resumen de la sentencia, penas aplicables, derechos del procesado y disposiciones finales]

NOTIFÍQUESE Y CÚMPLASE.

CONTEXTO LEGAL ({tipo}):
{contexto}

CASO A JUZGAR:
{pregunta}

IMPORTANTE: Basa tu sentencia ÚNICAMENTE en el {tipo} proporcionado. Cita artículos específicos."""
    else:
        return f"""Actúas como un JUEZ especializado en legislación ecuatoriana.

⚠️ IMPORTANTE: La información específica sobre este caso NO se encuentra en el documento legal proporcionado.

Estructura tu respuesta como sentencia, pero indicando la limitación:

📅 **FECHA Y HORA:**
[Fecha y hora actual de la sentencia]

⚖️ **RAZÓN DE LA SENTENCIA:**
No se cuenta con el marco legal específico en el documento proporcionado para este caso. Sin embargo, basándome en principios generales del derecho ecuatoriano:

🏛️ **VEREDICTO:**
SUSPENDIDO - Falta de marco legal específico en el documento proporcionado

🏢 **LUGAR DE RECLUSIÓN:**
No aplicable hasta contar con legislación específica

📋 **CONCLUSIÓN:**
Esta sentencia provisional se basa en conocimiento general jurídico y NO en el documento específico proporcionado.
Se requiere consultar la legislación correspondiente y asesoramiento legal especializado.

CASO PRESENTADO:
{pregunta}

⚠️ ADVERTENCIA JUDICIAL: Sentencia no definitiva por falta de marco legal específico."""

def guardar_en_excel(pregunta: str, respuesta: str, path: str = EXCEL_PATH):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws[get_column_letter(1) + "1"] = "Pregunta"
        ws[get_column_letter(2) + "1"] = "Respuesta"
        ws[get_column_letter(3) + "1"] = "tiempo"

    fila = ws.max_row + 1
    segundos_actuales = datetime.now().strftime("%S")
    ws[f"A{fila}"] = pregunta
    ws[f"B{fila}"] = respuesta
    ws[f"C{fila}"] = segundos_actuales
    wb.save(path)

# Función para insertar puntos en lotes para evitar timeouts
def insertar_puntos_en_lotes(chunks, vectores, tipo_documento, batch_size=BATCH_SIZE):
    """
    Inserta los puntos en Qdrant en lotes para evitar timeouts con documentos grandes
    """
    total_chunks = len(chunks)
    puntos_insertados = 0
    
    # Procesar en lotes
    for i in range(0, total_chunks, batch_size):
        end_idx = min(i + batch_size, total_chunks)
        batch_chunks = chunks[i:end_idx]
        batch_vectores = vectores[i:end_idx]
        
        # Crear puntos para este lote
        puntos_lote = [
            PointStruct(
                id=i + j, 
                vector=batch_vectores[j].tolist(), 
                payload={
                    "text": batch_chunks[j], 
                    "chunk_index": i + j,
                    "documento_tipo": tipo_documento.get("tipo", "Documento Legal"),
                    "documento_especialidad": tipo_documento.get("especialidad", "Derecho General")
                }
            )
            for j in range(len(batch_chunks))
        ]
        
        # Insertar lote con reintentos
        max_reintentos = 3
        for intento in range(max_reintentos):
            try:
                qdrant_client.upsert(collection_name=COLLECTION_NAME, points=puntos_lote)
                puntos_insertados += len(puntos_lote)
                print(f"Lote {i//batch_size + 1} insertado exitosamente. Progreso: {puntos_insertados}/{total_chunks}")
                break
            except Exception as e:
                if intento == max_reintentos - 1:
                    raise Exception(f"Error al insertar lote después de {max_reintentos} intentos: {str(e)}")
                print(f"Error en lote {i//batch_size + 1}, intento {intento + 1}: {str(e)}")
                import time
                time.sleep(2 ** intento)  # Backoff exponencial
    
    return puntos_insertados

# Función para detectar el tipo de documento legal
def detectar_tipo_documento(texto: str, filename: str = "") -> dict:
    """
    Detecta el tipo de documento legal basado en el contenido y nombre del archivo
    """
    texto_lower = texto.lower()
    filename_lower = filename.lower()
    
    # Patrones de identificación
    patrones = {
        "COIP": {
            "keywords": ["código orgánico integral penal", "coip", "delitos", "penas", "infracciones penales", "homicidio", "robo", "estafa"],
            "especialidad": "Derecho Penal",
            "descripcion": "Código Orgánico Integral Penal"
        },
        "Código de Comercio": {
            "keywords": ["código de comercio", "mercantil", "comerciante", "sociedad anónima", "contrato mercantil", "empresa"],
            "especialidad": "Derecho Mercantil",
            "descripcion": "Código de Comercio"
        },
        "Código de la Niñez": {
            "keywords": ["código de la niñez", "niños", "adolescentes", "menores", "patria potestad", "tutela"],
            "especialidad": "Derecho de Familia y Niñez",
            "descripcion": "Código de la Niñez y Adolescencia"
        },
        "Código Civil": {
            "keywords": ["código civil", "derecho civil", "personas", "bienes", "obligaciones", "contratos civiles"],
            "especialidad": "Derecho Civil",
            "descripcion": "Código Civil"
        },
        "Constitución": {
            "keywords": ["constitución", "derechos fundamentales", "garantías constitucionales", "estado", "poderes públicos"],
            "especialidad": "Derecho Constitucional",
            "descripcion": "Constitución"
        }
    }
    
    # Detectar por nombre de archivo primero
    for tipo, info in patrones.items():
        if any(keyword in filename_lower for keyword in info["keywords"][:2]):  # Solo los primeros 2 keywords más específicos
            return {
                "tipo": tipo,
                "especialidad": info["especialidad"],
                "descripcion": info["descripcion"],
                "confianza": "alta",
                "metodo": "filename"
            }
    
    # Detectar por contenido
    scores = {}
    for tipo, info in patrones.items():
        score = sum(1 for keyword in info["keywords"] if keyword in texto_lower)
        if score > 0:
            scores[tipo] = score
    
    if scores:
        tipo_detectado = max(scores, key=scores.get)
        max_score = scores[tipo_detectado]
        confianza = "alta" if max_score >= 3 else "media" if max_score >= 2 else "baja"
        
        return {
            "tipo": tipo_detectado,
            "especialidad": patrones[tipo_detectado]["especialidad"],
            "descripcion": patrones[tipo_detectado]["descripcion"],
            "confianza": confianza,
            "metodo": "contenido",
            "score": max_score
        }
    
    # Documento genérico si no se detecta
    return {
        "tipo": "Documento Legal Genérico",
        "especialidad": "Derecho General",
        "descripcion": "Documento Legal",
        "confianza": "baja",
        "metodo": "generico"
    }

# === Endpoints ===
# Verificar estado del servicio
@app.get("/status", summary="Verificar estado del servicio")
async def check_status():
    try:
        # Verificar conexión con Qdrant
        collections = qdrant_client.get_collections().collections
        collection_names = [collection.name for collection in collections]
        
        # Verificar API de OpenAI
        test_response = generar_respuesta_openai("Hola, di 'OK' si funcionas correctamente", "test")
        
        return {
            "estado": "ok",
            "qdrant_conectado": True,
            "openai_conectado": True,
            "colecciones_disponibles": collection_names,
            "test_openai": test_response,
            "version": "1.0.0"
        }
    except Exception as e:
        return {
            "estado": "error",
            "mensaje": str(e)
        }

# Subir documento PDF y cargar a Qdrant
@app.post("/documento/subir", summary="Subir documento PDF y cargar a Qdrant")
async def subir_documento(file: UploadFile = File(...)):
    try:
        # Validar tipo de archivo
        if not file.filename.endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Solo se aceptan archivos .pdf")

        # Validar tamaño del archivo (límite de 50MB)
        file_content = await file.read()
        if len(file_content) > 50 * 1024 * 1024:  # 50MB
            raise HTTPException(status_code=400, detail="El archivo es demasiado grande. Máximo 50MB.")

        # Guardar archivo
        ruta = os.path.join(UPLOAD_FOLDER, file.filename)
        with open(ruta, "wb") as f:
            f.write(file_content)

        # Procesar PDF y extraer texto
        try:
            chunks, vectores, tipo_documento = pdf_a_chunks(ruta)
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=f"Error al procesar PDF: {str(ve)}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error inesperado al procesar PDF: {str(e)}")

        # Validar que se extrajeron chunks
        if not chunks:
            raise HTTPException(status_code=400, detail="No se pudo extraer texto del PDF")

        # Actualizar información del documento actual
        global documento_actual
        documento_actual.update({
            "tipo": tipo_documento.get("tipo"),
            "especialidad": tipo_documento.get("especialidad"),
            "descripcion": tipo_documento.get("descripcion"),
            "filename": file.filename,
            "fecha_carga": datetime.now().isoformat()
        })

        # Inicializar colección de Qdrant
        try:
            inicializar_qdrant()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error al inicializar Qdrant: {str(e)}")

        # Insertar puntos en lotes para evitar timeouts
        try:
            puntos_insertados = insertar_puntos_en_lotes(chunks, vectores, tipo_documento)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error al cargar documento en Qdrant: {str(e)}")

        return {
            "estado": "ok",
            "fragmentos_cargados": puntos_insertados,
            "archivo": file.filename,
            "tamaño_archivo_mb": round(len(file_content) / (1024 * 1024), 2),
            "documento_detectado": {
                "tipo": tipo_documento.get("tipo"),
                "especialidad": tipo_documento.get("especialidad"),  
                "descripcion": tipo_documento.get("descripcion"),
                "confianza": tipo_documento.get("confianza"),
                "metodo_deteccion": tipo_documento.get("metodo")
            },
            "configuracion": {
                "chunk_size": CHUNK_SIZE,
                "overlap_size": OVERLAP_SIZE,
                "batch_size": BATCH_SIZE
            }
        }

    except HTTPException:
        # Re-lanzar excepciones HTTP
        raise
    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        print(f"Error detallado en subir_documento: {error_detalle}")
        raise HTTPException(status_code=500, detail=f"Error interno del servidor: {str(e)}")

class ConsultaChat(BaseModel):
    pregunta: str

@app.post("/chat", summary="Consulta al chatbot usando contexto de documentos")
async def consultar_chat(req: ConsultaChat):
    try:
        # Verificar si la colección existe
        collections = qdrant_client.get_collections().collections
        collection_names = [collection.name for collection in collections]
        
        if COLLECTION_NAME not in collection_names:
            raise HTTPException(
                status_code=404, 
                detail=f"La colección {COLLECTION_NAME} no existe. Por favor, sube un documento primero."
            )
            
        # Obtener el conteo de puntos en la colección
        collection_info = qdrant_client.get_collection(COLLECTION_NAME)
        if collection_info.points_count == 0:
            raise HTTPException(
                status_code=404, 
                detail=f"La colección {COLLECTION_NAME} está vacía. Por favor, sube un documento primero."
            )
            
        # Codificar la pregunta
        vector_pregunta = model_embeddings.encode([req.pregunta])[0]
        
        # Buscar contexto relevante con más resultados para documentos grandes
        try:
            resultados = qdrant_client.search(
                collection_name=COLLECTION_NAME,
                query_vector=vector_pregunta,
                limit=8,  # Aumentado para mejor cobertura de documentos grandes
                score_threshold=MIN_SIMILARITY_THRESHOLD
            )
        except Exception as e:
            print(f"Error al buscar en Qdrant: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error al buscar información: {str(e)}")
        
        # Verificar si hay resultados relevantes
        if not resultados or (resultados and resultados[0].score < MIN_SIMILARITY_THRESHOLD):
            # No hay contexto relevante, usar conocimiento general de IA
            prompt = construir_prompt("", req.pregunta, documento_actual, tiene_contexto_relevante=False)
            texto_respuesta = generar_respuesta_openai(prompt, req.pregunta, OPENAI_MODEL)
            
            # Guardar en Excel con indicación de respuesta basada en IA
            guardar_en_excel(f"[SIN CONTEXTO DOC] {req.pregunta}", texto_respuesta)
            
            return {"respuesta": texto_respuesta, "fuente": "conocimiento_ia", "modelo_usado": OPENAI_MODEL}
            
        # Hay contexto relevante, construir respuesta basada en documento
        contexto_combinado = []
        for resultado in resultados:
            contexto_combinado.append(f"[Relevancia: {resultado.score:.3f}] {resultado.payload['text']}")
        
        contexto = "\n\n".join(contexto_combinado)
        prompt = construir_prompt(contexto, req.pregunta, documento_actual, tiene_contexto_relevante=True)

        # Generar respuesta
        texto_respuesta = generar_respuesta_openai(prompt, req.pregunta, OPENAI_MODEL)

        # Agregar información sobre las fuentes consultadas
        num_fragmentos = len(resultados)
        max_score = max([r.score for r in resultados])
        min_score = min([r.score for r in resultados])
        
        info_fuentes = f"\n\n📚 **Información de consulta:**\n- Documento: {documento_actual.get('descripcion', 'Documento Legal')}\n- Especialidad: {documento_actual.get('especialidad', 'Derecho General')}\n- Fragmentos consultados: {num_fragmentos}\n- Relevancia máxima: {max_score:.3f}\n- Relevancia mínima: {min_score:.3f}"
        texto_respuesta += info_fuentes

        # Guardar pregunta y respuesta en Excel
        guardar_en_excel(req.pregunta, texto_respuesta)

        return {
            "respuesta": texto_respuesta, 
            "fuente": "documento",
            "documento_tipo": documento_actual.get('tipo'),
            "documento_especialidad": documento_actual.get('especialidad'),
            "fragmentos_consultados": num_fragmentos,
            "relevancia_maxima": max_score,
            "modelo_usado": OPENAI_MODEL
        }

    except HTTPException as he:
        # Re-lanzar excepciones HTTP
        raise he
    except Exception as e:
        import traceback
        error_detalle = traceback.format_exc()
        print(f"Error detallado: {error_detalle}")
        raise HTTPException(status_code=500, detail=f"Error al generar respuesta: {str(e)}")

# Obtener estadísticas del documento cargado
@app.get("/documento/estadisticas", summary="Obtener estadísticas del documento cargado")
async def obtener_estadisticas_documento():
    try:
        # Verificar si la colección existe
        collections = qdrant_client.get_collections().collections
        collection_names = [collection.name for collection in collections]
        
        if COLLECTION_NAME not in collection_names:
            return {
                "estado": "sin_documento",
                "mensaje": "No hay documento cargado"
            }
            
        # Obtener información de la colección
        collection_info = qdrant_client.get_collection(COLLECTION_NAME)
        
        if collection_info.points_count == 0:
            return {
                "estado": "coleccion_vacia",
                "mensaje": "La colección existe pero está vacía"
            }
        
        # Obtener algunos puntos de muestra para estadísticas
        puntos_muestra = qdrant_client.scroll(
            collection_name=COLLECTION_NAME,
            limit=10,
            with_payload=True
        )[0]
        
        # Calcular estadísticas básicas
        longitudes_texto = [len(punto.payload.get("text", "")) for punto in puntos_muestra]
        longitud_promedio = sum(longitudes_texto) / len(longitudes_texto) if longitudes_texto else 0
        
        return {
            "estado": "documento_cargado",
            "total_fragmentos": collection_info.points_count,
            "longitud_promedio_fragmento": round(longitud_promedio, 2),
            "longitud_minima_muestra": min(longitudes_texto) if longitudes_texto else 0,
            "longitud_maxima_muestra": max(longitudes_texto) if longitudes_texto else 0,
            "documento_actual": {
                "tipo": documento_actual.get('tipo'),
                "especialidad": documento_actual.get('especialidad'),
                "descripcion": documento_actual.get('descripcion'),
                "filename": documento_actual.get('filename'),
                "fecha_carga": documento_actual.get('fecha_carga')
            },
            "configuracion": {
                "chunk_size": CHUNK_SIZE,
                "overlap_size": OVERLAP_SIZE,
                "umbral_similitud": MIN_SIMILARITY_THRESHOLD
            }
        }
        
    except Exception as e:
        return {
            "estado": "error",
            "mensaje": f"Error al obtener estadísticas: {str(e)}"
        }

# Limpiar colección (útil para resolver problemas de carga)
@app.delete("/documento/limpiar", summary="Limpiar colección de documentos")
async def limpiar_coleccion():
    try:
        collections = qdrant_client.get_collections().collections
        collection_names = [collection.name for collection in collections]
        
        if COLLECTION_NAME in collection_names:
            qdrant_client.delete_collection(COLLECTION_NAME)
            return {
                "estado": "ok",
                "mensaje": f"Colección {COLLECTION_NAME} eliminada exitosamente"
            }
        else:
            return {
                "estado": "info",
                "mensaje": "No hay colección para limpiar"
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al limpiar colección: {str(e)}")

# Verificar conectividad con Qdrant
@app.get("/qdrant/test", summary="Verificar conectividad con Qdrant")
async def test_qdrant():
    try:
        # Test básico de conectividad
        collections = qdrant_client.get_collections()
        
        # Test de creación temporal
        test_collection = "test_connection"
        try:
            qdrant_client.recreate_collection(
                collection_name=test_collection,
                vectors_config=VectorParams(size=10, distance=Distance.COSINE)
            )
            qdrant_client.delete_collection(test_collection)
            conectividad_completa = True
            mensaje_test = "Qdrant funcionando correctamente"
        except Exception as test_error:
            conectividad_completa = False
            mensaje_test = f"Conectado pero error en operaciones: {str(test_error)}"
        
        return {
            "estado": "ok" if conectividad_completa else "warning",
            "conectividad": "completa" if conectividad_completa else "parcial",
            "colecciones_disponibles": len(collections.collections),
            "mensaje": mensaje_test
        }
        
    except Exception as e:
        return {
            "estado": "error",
            "conectividad": "fallida",
            "mensaje": f"Error de conectividad: {str(e)}"
        }

# Obtener información del documento actual
@app.get("/documento/info", summary="Obtener información del documento actualmente cargado")
async def obtener_info_documento():
    if not documento_actual.get('tipo'):
        return {
            "estado": "sin_documento",
            "mensaje": "No hay documento cargado actualmente"
        }
    
    return {
        "estado": "documento_disponible",
        "documento": documento_actual
    }

# Función para generar respuestas con OpenAI
def generar_respuesta_openai(prompt: str, pregunta: str = "", modelo: str = "gpt-3.5-turbo") -> str:
    """
    Genera una respuesta usando la API de OpenAI con formato de sentencia
    Modelos disponibles: gpt-3.5-turbo, gpt-4, gpt-4-turbo-preview
    """
    try:
        client = openai.OpenAI(api_key=config["OPENAI_API_KEY"])
        
        # Configuraciones según el modelo
        max_tokens = 3000 if "gpt-4" in modelo else 2000
        
        response = client.chat.completions.create(
            model=modelo,
            messages=[
                {
                    "role": "system", 
                    "content": "Eres un juez especializado en derecho ecuatoriano. Siempre debes responder con el formato estructurado de una sentencia judicial, incluyendo fecha, razón, veredicto, lugar de reclusión y conclusión. Sé preciso en las citas legales y mantén la imparcialidad judicial."
                },
                {
                    "role": "user", 
                    "content": prompt
                }
            ],
            max_tokens=max_tokens,
            temperature=0.2,  # Más determinista para respuestas judiciales
            top_p=0.95,
            frequency_penalty=0.0,
            presence_penalty=0.0
        )
        
        respuesta = response.choices[0].message.content.strip()
        
        # Post-procesar para asegurar formato de sentencia
        respuesta_formateada = post_procesar_sentencia(respuesta, pregunta)
        
        return respuesta_formateada
        
    except openai.AuthenticationError:
        error_msg = "Error de autenticación con OpenAI. Verifica tu API key."
    except openai.RateLimitError:
        error_msg = "Límite de velocidad excedido en OpenAI. Intenta de nuevo en unos momentos."
    except openai.APIError as e:
        error_msg = f"Error de API de OpenAI: {str(e)}"
    except Exception as e:
        error_msg = f"Error al generar respuesta con OpenAI: {str(e)}"
    
    # Respuesta de error también en formato de sentencia
    fecha_sentencia = generar_fecha_sentencia()
    return f"""
📅 **FECHA Y HORA:**
{fecha_sentencia}

⚖️ **RAZÓN DE LA SENTENCIA:**
Error técnico en el sistema de procesamiento legal: {error_msg}

🏛️ **VEREDICTO:**
SUSPENDIDO - Error en el sistema

🏢 **LUGAR DE RECLUSIÓN:**
No aplicable

📋 **CONCLUSIÓN:**
Se requiere revisión técnica del sistema antes de proceder con el análisis jurídico.

NOTIFÍQUESE AL ADMINISTRADOR DEL SISTEMA.
"""

# Configurar modelo de OpenAI
class ModeloConfig(BaseModel):
    modelo: str

@app.post("/configuracion/modelo", summary="Configurar modelo de OpenAI a usar")
async def configurar_modelo(config: ModeloConfig):
    global OPENAI_MODEL
    
    modelos_disponibles = [
        "gpt-3.5-turbo",
        "gpt-3.5-turbo-16k", 
        "gpt-4",
        "gpt-4-turbo-preview",
        "gpt-4o",
        "gpt-4o-mini"
    ]
    
    if config.modelo not in modelos_disponibles:
        raise HTTPException(
            status_code=400, 
            detail=f"Modelo no válido. Modelos disponibles: {', '.join(modelos_disponibles)}"
        )
    
    # Verificar que el modelo funciona con una consulta de prueba
    try:
        test_response = generar_respuesta_openai("Di 'OK' si funcionas", "test", config.modelo)
        OPENAI_MODEL = config.modelo
        
        return {
            "estado": "ok",
            "modelo_anterior": OPENAI_MODEL,
            "modelo_actual": config.modelo,
            "test_response": test_response,
            "mensaje": f"Modelo cambiado exitosamente a {config.modelo}"
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al probar el modelo {config.modelo}: {str(e)}"
        )

@app.get("/configuracion/modelo", summary="Obtener modelo actual de OpenAI")
async def obtener_modelo_actual():
    return {
        "modelo_actual": OPENAI_MODEL,
        "modelos_disponibles": [
            "gpt-3.5-turbo",
            "gpt-3.5-turbo-16k", 
            "gpt-4",
            "gpt-4-turbo-preview",
            "gpt-4o",
            "gpt-4o-mini"
        ]
    }

# === Funciones para formato de sentencia ===
def generar_fecha_sentencia():
    """
    Genera la fecha y hora actual para la sentencia en formato judicial
    """
    # Zona horaria de Ecuador
    tz_ecuador = pytz.timezone('America/Guayaquil')
    ahora = datetime.now(tz_ecuador)
    
    return ahora.strftime("%d de %B de %Y, %H:%M horas")

def obtener_tipo_carcel(delito: str) -> str:
    """
    Determina el tipo de centro penitenciario según el delito
    """
    tipos_carcel = {
        "homicidio": "Centro de Rehabilitación Social de Máxima Seguridad",
        "asesinato": "Centro de Rehabilitación Social de Máxima Seguridad", 
        "robo": "Centro de Rehabilitación Social de Mediana Seguridad",
        "hurto": "Centro de Rehabilitación Social de Mínima Seguridad",
        "estafa": "Centro de Rehabilitación Social de Mediana Seguridad",
        "violación": "Centro de Rehabilitación Social de Máxima Seguridad",
        "secuestro": "Centro de Rehabilitación Social de Máxima Seguridad",
        "narcotráfico": "Centro de Rehabilitación Social de Máxima Seguridad",
        "lavado de activos": "Centro de Rehabilitación Social de Mediana Seguridad",
        "femicidio": "Centro de Rehabilitación Social de Máxima Seguridad",
        "violencia intrafamiliar": "Centro de Rehabilitación Social de Mediana Seguridad"
    }
    
    # Buscar por palabras clave en el delito
    delito_lower = delito.lower()
    for key, carcel in tipos_carcel.items():
        if key in delito_lower:
            return carcel
    
    # Por defecto
    return "Centro de Rehabilitación Social de Mediana Seguridad"

def post_procesar_sentencia(respuesta: str, pregunta: str) -> str:
    """
    Post-procesa la respuesta para asegurar el formato de sentencia
    """
    # Reemplazar placeholder de fecha si existe
    fecha_sentencia = generar_fecha_sentencia()
    respuesta = respuesta.replace("[Fecha y hora actual de la sentencia]", fecha_sentencia)
    
    # Si la respuesta no tiene el formato completo, estructurarla
    if not all(seccion in respuesta for seccion in ["📅 **FECHA Y HORA:**", "⚖️ **RAZÓN DE LA SENTENCIA:**", "🏛️ **VEREDICTO:**"]):
        # Restructurar la respuesta en formato de sentencia
        respuesta_estructurada = f"""
📅 **FECHA Y HORA:**
{fecha_sentencia}

⚖️ **RAZÓN DE LA SENTENCIA:**
{respuesta}

🏛️ **VEREDICTO:**
Pendiente de análisis legal específico basado en el contexto proporcionado

🏢 **LUGAR DE RECLUSIÓN:**
{obtener_tipo_carcel(pregunta)}

📋 **CONCLUSIÓN:**
Se requiere análisis legal más detallado para determinar sentencia definitiva.
Sentencia sujeta a los procedimientos establecidos en la legislación ecuatoriana.

NOTIFÍQUESE Y CÚMPLASE.
"""
        return respuesta_estructurada
    
    return respuesta

# Endpoint para generar sentencia de ejemplo
@app.post("/sentencia/ejemplo", summary="Generar ejemplo de sentencia judicial")
async def generar_sentencia_ejemplo():
    """
    Endpoint para generar una sentencia de ejemplo con el formato correcto
    """
    fecha_sentencia = generar_fecha_sentencia()
    
    ejemplo_sentencia = f"""
📅 **FECHA Y HORA:**
{fecha_sentencia}

⚖️ **RAZÓN DE LA SENTENCIA:**
El acusado Juan Carlos Pérez Mendoza, identificado con cédula 1234567890, ha sido procesado por el delito de hurto tipificado en el artículo 196 del COIP. Los elementos probatorios presentados durante el juicio oral demuestran de manera fehaciente que el acusado sustrajo bienes muebles ajenos valorados en $500 USD del local comercial "La Esperanza", sin emplear violencia, intimidación o fuerza en las personas. La materialidad del delito se encuentra acreditada con el testimonio de la víctima, las grabaciones de seguridad y el reconocimiento en rueda de personas.

🏛️ **VEREDICTO:**
CULPABLE del delito de hurto simple previsto y sancionado en el artículo 196 del Código Orgánico Integral Penal, en concordancia con los artículos 234 y siguientes del mismo cuerpo legal.

🏢 **LUGAR DE RECLUSIÓN:**
Centro de Rehabilitación Social de Mínima Seguridad "La Roca" por el término de 6 meses, con posibilidad de acogerse a los beneficios penitenciarios establecidos en el régimen progresivo según la Ley del Sistema Nacional de Rehabilitación Social.

📋 **CONCLUSIÓN:**
Se condena al procesado Juan Carlos Pérez Mendoza a 6 meses de privación de libertad y al pago de una multa equivalente al doble del valor de los bienes sustraídos ($1,000 USD). El sentenciado deberá reparar integralmente el daño causado a la víctima. La presente sentencia es apelable ante la Corte Provincial dentro del término de 10 días contados desde su notificación, conforme lo establecido en el Código Orgánico de la Función Judicial.

NOTIFÍQUESE Y CÚMPLASE.

Dra. María Elena Rodríguez
JUEZA DE GARANTÍAS PENALES
"""
    
    return {
        "sentencia_ejemplo": ejemplo_sentencia,
        "formato_aplicado": "✅ Formato de sentencia judicial ecuatoriana",
        "fecha_generacion": fecha_sentencia,
        "caracteristicas": {
            "estructura": "Fecha, Razón, Veredicto, Reclusión, Conclusión",
            "tipo_caso": "Hurto simple (Artículo 196 COIP)",
            "centro_penitenciario": "Mínima Seguridad",
            "tiempo_condena": "6 meses"
        }
    }
